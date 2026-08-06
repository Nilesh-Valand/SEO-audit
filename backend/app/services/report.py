from __future__ import annotations

import base64
import csv
import io
import os
import tempfile
from collections import defaultdict
from datetime import datetime
from typing import Any

from sqlalchemy import select

from app.db.database import SessionLocal
from app.models import CrawlRun, CrawlRunScore, Project
from app.rules.engine import RuleEngine
from app.rules.recommendations import recommendation_for_rule
from app.services.issues import load_issue_views

# Same weights as RuleEngine scoring: impact = weight × pages affected.
SEVERITY_WEIGHTS = {
    "critical": 10,
    "high": 5,
    "medium": 2,
    "low": 1,
}


class ReportService:
    def __init__(self) -> None:
        self._rules_by_id = {rule.id: rule for rule in RuleEngine().rules}

    def build_report(self, crawl_run_id: int) -> dict[str, Any]:
        with SessionLocal() as db:
            crawl_run = db.get(CrawlRun, crawl_run_id)
            if crawl_run is None:
                raise ValueError(f"Crawl run {crawl_run_id} not found.")

            project = db.get(Project, crawl_run.project_id)
            scores = db.scalars(
                select(CrawlRunScore)
                .where(CrawlRunScore.crawl_id == crawl_run_id)
                .order_by(CrawlRunScore.category.asc())
            ).all()
            issues = load_issue_views(db, crawl_run_id, rules_by_id=self._rules_by_id)
            issues.sort(key=lambda item: (item.category, item.id))

            category_scores = {
                score.category: score.score for score in scores if score.category != "overall"
            }
            overall_score = next(
                (score.score for score in scores if score.category == "overall"), None
            )

            severity_counts = {severity: 0 for severity in ("critical", "high", "medium", "low")}
            category_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
            recommendation_groups: dict[tuple[str, str, str], dict[str, Any]] = {}

            # Site issues: one row each from site_issues — never fan out per page.
            site_issues: list[dict[str, Any]] = []
            # Page issues: group only that URL's page_issues rows.
            page_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)

            for issue in issues:
                severity_counts[issue.severity] = severity_counts.get(issue.severity, 0) + 1
                is_site = issue.scope != "page"

                payload = {
                    "id": issue.id,
                    "rule": issue.check_name,
                    "severity": issue.severity,
                    "category": issue.category,
                    "message": issue.details,
                    "scope": issue.scope,
                }

                if is_site:
                    site_issues.append(payload)
                    # Category inventory keeps site findings URL-less (no per-page dupes).
                    category_groups[issue.category].append(
                        {
                            "url": None,
                            "rule": issue.check_name,
                            "severity": issue.severity,
                            "message": issue.details,
                            "scope": issue.scope,
                        }
                    )
                else:
                    url = issue.url or ""
                    page_groups[url].append(payload)
                    category_groups[issue.category].append(
                        {
                            "url": issue.url,
                            "rule": issue.check_name,
                            "severity": issue.severity,
                            "message": issue.details,
                            "scope": "page",
                        }
                    )

                recommendation_key = (issue.check_name, issue.severity, issue.category)
                if recommendation_key not in recommendation_groups:
                    recommendation_groups[recommendation_key] = {
                        "rule": issue.check_name,
                        "severity": issue.severity,
                        "category": issue.category,
                        "message": recommendation_for_rule(issue.check_name, issue.details),
                        "pages_affected": 0,
                    }
                # Site findings count once; page findings count once per affected URL row.
                recommendation_groups[recommendation_key]["pages_affected"] += 1

            page_issues = [
                {
                    "url": url,
                    "issue_count": len(rows),
                    "issues": rows,
                }
                for url, rows in sorted(page_groups.items(), key=lambda item: item[0])
            ]

            site_issue_count = len(site_issues)
            pages_with_issues = len(page_issues)
            page_issue_count = sum(group["issue_count"] for group in page_issues)
            summary_text = (
                f"{site_issue_count} site issue{'s' if site_issue_count != 1 else ''}, "
                f"{pages_with_issues} page{'s' if pages_with_issues != 1 else ''} with issues, "
                f"{page_issue_count} total page-level finding"
                f"{'s' if page_issue_count != 1 else ''}."
            )

            categories = [
                {
                    "name": self._labelize(category),
                    "score": category_scores.get(category),
                    "issues": issues_for_category,
                }
                for category, issues_for_category in sorted(category_groups.items())
            ]

            recommendations = sorted(
                recommendation_groups.values(),
                key=lambda item: (
                    -(
                        SEVERITY_WEIGHTS.get(item["severity"], 1)
                        * item["pages_affected"]
                    ),
                    item["rule"],
                ),
            )

            return {
                "project": {
                    "id": project.id if project else crawl_run.project_id,
                    "domain": project.domain if project else None,
                },
                "crawl_date": (
                    crawl_run.finished_at.isoformat()
                    if crawl_run.finished_at
                    else crawl_run.started_at.isoformat()
                    if crawl_run.started_at
                    else None
                ),
                "overall_score": overall_score,
                "category_scores": category_scores,
                "summary": {
                    "total_pages": crawl_run.total_urls,
                    "total_issues": site_issue_count + page_issue_count,
                    "site_issue_count": site_issue_count,
                    "pages_with_issues": pages_with_issues,
                    "page_issue_count": page_issue_count,
                    "summary_text": summary_text,
                    "issues_by_severity": severity_counts,
                },
                "site_issues": site_issues,
                "page_issues": page_issues,
                "categories": categories,
                "recommendations": recommendations,
            }

    def iter_csv_rows(self, crawl_run_id: int):
        with SessionLocal() as db:
            crawl_run = db.get(CrawlRun, crawl_run_id)
            if crawl_run is None:
                raise ValueError(f"Crawl run {crawl_run_id} not found.")

            buffer = io.StringIO()
            writer = csv.writer(buffer)
            writer.writerow(["section", "url", "category", "rule", "severity", "message", "scope"])
            yield buffer.getvalue()
            buffer.seek(0)
            buffer.truncate(0)

            issues = load_issue_views(db, crawl_run_id, rules_by_id=self._rules_by_id)
            issues.sort(
                key=lambda item: (
                    0 if item.scope != "page" else 1,
                    item.url or "",
                    item.category,
                    item.id,
                )
            )
            for issue in issues:
                is_site = issue.scope != "page"
                writer.writerow(
                    [
                        "site" if is_site else "page",
                        "" if is_site else (issue.url or ""),
                        issue.category,
                        issue.check_name,
                        issue.severity,
                        issue.details,
                        issue.scope,
                    ]
                )
                yield buffer.getvalue()
                buffer.seek(0)
                buffer.truncate(0)

    def generate_pdf_file(self, crawl_run_id: int) -> tuple[str, str]:
        """Build a PDF report using ReportLab (Windows-friendly; no GTK/WeasyPrint)."""
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            Image,
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        report = self.build_report(crawl_run_id)
        chart_paths = self._build_chart_files(report)

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        temp_file.close()
        filename = f"crawl-run-{crawl_run_id}-report.pdf"

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CoverTitle",
            parent=styles["Title"],
            fontSize=22,
            spaceAfter=18,
            alignment=TA_CENTER,
        )
        score_style = ParagraphStyle(
            "CoverScore",
            parent=styles["Title"],
            fontSize=42,
            textColor=colors.HexColor("#0284c7"),
            alignment=TA_CENTER,
            spaceAfter=12,
        )
        heading_style = ParagraphStyle(
            "SectionHeading",
            parent=styles["Heading2"],
            fontSize=14,
            spaceBefore=16,
            spaceAfter=8,
            textColor=colors.HexColor("#111827"),
        )
        body_style = ParagraphStyle(
            "Body",
            parent=styles["Normal"],
            fontSize=9,
            leading=12,
        )
        small_style = ParagraphStyle(
            "Small",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#4b5563"),
        )

        def safe(text: Any) -> str:
            return (
                str(text if text is not None else "")
                .replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;")
            )

        story: list[Any] = []
        domain = report["project"].get("domain") or "SEO Audit Report"
        overall = report["overall_score"]
        overall_display = f"{overall:.0f}" if isinstance(overall, (int, float)) else "--"

        story.append(Spacer(1, 1.8 * inch))
        story.append(Paragraph(safe(domain), title_style))
        story.append(Paragraph(overall_display, score_style))
        story.append(
            Paragraph(
                f"Generated {datetime.utcnow().isoformat()} · Crawl date: {safe(report.get('crawl_date') or 'N/A')}",
                ParagraphStyle("Meta", parent=small_style, alignment=TA_CENTER),
            )
        )
        story.append(PageBreak())

        story.append(Paragraph("Executive Summary", heading_style))
        summary = report["summary"]
        severity = summary.get("issues_by_severity") or {}
        summary_data = [
            ["Total pages", str(summary.get("total_pages", 0))],
            [
                "Issue summary",
                str(
                    summary.get("summary_text")
                    or (
                        f"{summary.get('site_issue_count', 0)} site issues, "
                        f"{summary.get('pages_with_issues', 0)} pages with issues, "
                        f"{summary.get('page_issue_count', 0)} total page-level findings."
                    )
                ),
            ],
            ["Site issues", str(summary.get("site_issue_count", 0))],
            ["Pages with issues", str(summary.get("pages_with_issues", 0))],
            ["Page-level findings", str(summary.get("page_issue_count", 0))],
            ["Critical", str(severity.get("critical", 0))],
            ["High", str(severity.get("high", 0))],
            ["Medium", str(severity.get("medium", 0))],
            ["Low", str(severity.get("low", 0))],
        ]
        summary_table = Table(summary_data, colWidths=[2.2 * inch, 4.3 * inch])
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f3f4f6")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.append(summary_table)

        story.append(Paragraph("Category Scores", heading_style))
        category_rows = [["Category", "Score"]]
        for category, score in sorted((report.get("category_scores") or {}).items()):
            category_rows.append([self._labelize(category), f"{score:.2f}"])
        if len(category_rows) == 1:
            category_rows.append(["—", "—"])
        cat_table = Table(category_rows, colWidths=[4.0 * inch, 2.5 * inch])
        cat_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0284c7")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        story.append(cat_table)

        if chart_paths.get("category_scores"):
            story.append(Spacer(1, 0.2 * inch))
            story.append(Image(chart_paths["category_scores"], width=6.5 * inch, height=2.4 * inch))
        if chart_paths.get("severity"):
            story.append(Spacer(1, 0.15 * inch))
            story.append(Image(chart_paths["severity"], width=6.5 * inch, height=2.4 * inch))

        story.append(Paragraph("Prioritized Recommendations", heading_style))
        recommendations = report.get("recommendations") or []
        if not recommendations:
            story.append(Paragraph("No recommendations for this run.", body_style))
        else:
            for index, item in enumerate(recommendations[:40], start=1):
                story.append(
                    Paragraph(
                        f"<b>{index}. {safe(item.get('rule'))}</b> "
                        f"({safe(item.get('severity'))}, {safe(item.get('pages_affected'))} pages) — "
                        f"{safe(item.get('message'))}",
                        body_style,
                    )
                )
                story.append(Spacer(1, 4))

        # --- Site Issues (one row each; never duplicated per page) ---
        story.append(PageBreak())
        story.append(Paragraph("Site Issues", heading_style))
        story.append(
            Paragraph(
                "One row per site-level finding (robots, sitemap, cross-page, homepage). "
                "These are not repeated under individual pages.",
                body_style,
            )
        )
        site_issues = report.get("site_issues") or []
        if not site_issues:
            story.append(Paragraph("No site issues for this run.", body_style))
        else:
            rows = [["Severity", "Scope", "Rule", "Message"]]
            for issue in site_issues[:300]:
                rows.append(
                    [
                        Paragraph(safe(issue.get("severity")), small_style),
                        Paragraph(safe(issue.get("scope")), small_style),
                        Paragraph(safe(issue.get("rule")), small_style),
                        Paragraph(safe(issue.get("message")), small_style),
                    ]
                )
            site_table = Table(
                rows, colWidths=[0.85 * inch, 0.95 * inch, 1.4 * inch, 3.3 * inch]
            )
            site_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e5e7eb")),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 8),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 4),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.white, colors.HexColor("#f9fafb")],
                        ),
                    ]
                )
            )
            story.append(site_table)

        # --- Page Issues (grouped by URL) ---
        story.append(PageBreak())
        story.append(Paragraph("Page Issues", heading_style))
        story.append(
            Paragraph(
                "Findings from page_issues, grouped by URL. Only that page's rows are listed.",
                body_style,
            )
        )
        page_groups = report.get("page_issues") or []
        if not page_groups:
            story.append(Paragraph("No page-level issues for this run.", body_style))
        else:
            for group in page_groups[:150]:
                url = safe(group.get("url") or "—")
                count = group.get("issue_count") or len(group.get("issues") or [])
                story.append(
                    Paragraph(f"<b>{url}</b> — {count} finding(s)", body_style)
                )
                rows = [["Severity", "Rule", "Message"]]
                for issue in (group.get("issues") or [])[:80]:
                    rows.append(
                        [
                            Paragraph(safe(issue.get("severity")), small_style),
                            Paragraph(safe(issue.get("rule")), small_style),
                            Paragraph(safe(issue.get("message")), small_style),
                        ]
                    )
                page_table = Table(
                    rows, colWidths=[0.85 * inch, 1.5 * inch, 4.15 * inch]
                )
                page_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e5e7eb")),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, 0), 8),
                            ("VALIGN", (0, 0), (-1, -1), "TOP"),
                            ("LEFTPADDING", (0, 0), (-1, -1), 4),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                            ("TOPPADDING", (0, 0), (-1, -1), 4),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                            (
                                "ROWBACKGROUNDS",
                                (0, 1),
                                (-1, -1),
                                [colors.white, colors.HexColor("#f9fafb")],
                            ),
                        ]
                    )
                )
                story.append(page_table)
                story.append(Spacer(1, 8))

        doc = SimpleDocTemplate(
            temp_file.name,
            pagesize=letter,
            leftMargin=0.6 * inch,
            rightMargin=0.6 * inch,
            topMargin=0.6 * inch,
            bottomMargin=0.6 * inch,
            title=f"SEO Audit — {domain}",
        )
        try:
            doc.build(story)
        finally:
            for path in chart_paths.values():
                remove_file(path)

        return temp_file.name, filename

    def _build_chart_files(self, report: dict[str, Any]) -> dict[str, str]:
        import matplotlib.pyplot as plt

        paths: dict[str, str] = {}
        category_scores = report.get("category_scores") or {}
        if category_scores:
            labels = list(category_scores.keys())
            values = [category_scores[key] for key in labels]
            paths["category_scores"] = self._chart_to_temp_png(
                plt, labels, values, "Category Scores"
            )

        severity_counts = (report.get("summary") or {}).get("issues_by_severity") or {}
        if severity_counts:
            labels = list(severity_counts.keys())
            values = [severity_counts[key] for key in labels]
            paths["severity"] = self._chart_to_temp_png(plt, labels, values, "Issues by Severity")
        return paths

    def _chart_to_temp_png(self, plt: Any, labels: list[str], values: list[Any], title: str) -> str:
        fig, ax = plt.subplots(figsize=(8, 3.2))
        ax.bar([self._labelize(str(label)) for label in labels], values, color="#0284c7")
        ax.set_title(title)
        ax.grid(axis="y", linestyle="--", alpha=0.3)
        fig.tight_layout()
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        temp_file.close()
        fig.savefig(temp_file.name, format="png", dpi=140)
        plt.close(fig)
        return temp_file.name

    def _build_chart_images(self, report: dict[str, Any]) -> dict[str, str]:
        import matplotlib.pyplot as plt

        category_chart = self._chart_to_data_uri(
            plt,
            list(report["category_scores"].keys()),
            [report["category_scores"][key] for key in report["category_scores"]],
            "Category Scores",
        )
        severity_counts = report["summary"]["issues_by_severity"]
        severity_chart = self._chart_to_data_uri(
            plt,
            list(severity_counts.keys()),
            [severity_counts[key] for key in severity_counts],
            "Issues by Severity",
        )
        return {"category_scores": category_chart, "severity": severity_chart}

    def generate_xlsx_file(self, crawl_run_id: int) -> tuple[str, str]:
        from openpyxl import Workbook

        report = self.build_report(crawl_run_id)
        workbook = Workbook(write_only=True)

        summary_sheet = workbook.create_sheet("Summary")
        summary_sheet.append(["Project", report["project"]["domain"] or ""])
        summary_sheet.append(["Crawl Date", report["crawl_date"] or ""])
        summary_sheet.append(["Overall Score", report["overall_score"] or ""])
        summary_sheet.append(
            ["Issue Summary", (report.get("summary") or {}).get("summary_text") or ""]
        )
        summary_sheet.append(
            ["Site Issues", (report.get("summary") or {}).get("site_issue_count", 0)]
        )
        summary_sheet.append(
            ["Pages With Issues", (report.get("summary") or {}).get("pages_with_issues", 0)]
        )
        summary_sheet.append(
            [
                "Page-Level Findings",
                (report.get("summary") or {}).get("page_issue_count", 0),
            ]
        )
        summary_sheet.append([])
        summary_sheet.append(["Category", "Score"])
        for category, score in report["category_scores"].items():
            summary_sheet.append([self._labelize(category), score])

        site_sheet = workbook.create_sheet("Site Issues")
        site_sheet.append(["scope", "category", "rule", "severity", "message"])
        for issue in report.get("site_issues") or []:
            site_sheet.append(
                [
                    issue.get("scope") or "site",
                    issue.get("category") or "",
                    issue.get("rule") or "",
                    issue.get("severity") or "",
                    issue.get("message") or "",
                ]
            )

        page_sheet = workbook.create_sheet("Page Issues")
        page_sheet.append(["url", "category", "rule", "severity", "message"])
        for group in report.get("page_issues") or []:
            url = group.get("url") or ""
            for issue in group.get("issues") or []:
                page_sheet.append(
                    [
                        url,
                        issue.get("category") or "",
                        issue.get("rule") or "",
                        issue.get("severity") or "",
                        issue.get("message") or "",
                    ]
                )

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            workbook.save(temp_file.name)
            return temp_file.name, f"crawl-run-{crawl_run_id}-report.xlsx"

    def _build_chart_images(self, report: dict[str, Any]) -> dict[str, str]:
        import matplotlib.pyplot as plt

        category_chart = self._chart_to_data_uri(
            plt,
            list(report["category_scores"].keys()),
            [report["category_scores"][key] for key in report["category_scores"]],
            "Category Scores",
        )
        severity_counts = report["summary"]["issues_by_severity"]
        severity_chart = self._chart_to_data_uri(
            plt,
            list(severity_counts.keys()),
            [severity_counts[key] for key in severity_counts],
            "Issues by Severity",
        )
        return {"category_scores": category_chart, "severity": severity_chart}

    def _chart_to_data_uri(self, plt: Any, labels: list[str], values: list[Any], title: str) -> str:
        fig, ax = plt.subplots(figsize=(8, 3.5))
        ax.bar(labels, values, color="#0284c7")
        ax.set_title(title)
        ax.grid(axis="y", linestyle="--", alpha=0.3)
        fig.tight_layout()
        buffer = io.BytesIO()
        fig.savefig(buffer, format="png", dpi=150)
        plt.close(fig)
        encoded = base64.b64encode(buffer.getvalue()).decode("utf-8")
        return f"data:image/png;base64,{encoded}"

    def _sheet_name(self, category: str) -> str:
        return self._labelize(category)[:31]

    def _labelize(self, value: str) -> str:
        return value.replace("_", " ").title()

    def _pdf_template(self) -> str:
        return """
        <html>
          <head>
            <style>
              body { font-family: Arial, sans-serif; color: #111827; font-size: 12px; }
              .cover { page-break-after: always; text-align: center; padding-top: 120px; }
              .score { font-size: 56px; font-weight: bold; color: #0284c7; }
              .section { margin-top: 28px; }
              .grid { display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }
              .card { border: 1px solid #e5e7eb; border-radius: 10px; padding: 16px; margin-bottom: 18px; }
              table { width: 100%; border-collapse: collapse; margin-top: 10px; }
              th, td { border: 1px solid #e5e7eb; padding: 8px; vertical-align: top; text-align: left; }
              th { background: #f9fafb; }
              h1, h2, h3 { margin: 0 0 12px 0; }
              img { width: 100%; border: 1px solid #e5e7eb; border-radius: 8px; }
              ul { padding-left: 20px; }
            </style>
          </head>
          <body>
            <div class="cover">
              <h1>{{ report.project.domain or "SEO Audit Report" }}</h1>
              <div class="score">{{ report.overall_score if report.overall_score is not none else "--" }}</div>
              <p>Generated {{ generated_at }}</p>
            </div>

            <div class="section">
              <h2>Executive Summary</h2>
              <div class="grid">
                <div class="card">
                  <p><strong>Total Pages:</strong> {{ report.summary.total_pages }}</p>
                  <p><strong>Summary:</strong> {{ report.summary.summary_text }}</p>
                  <p><strong>Crawl Date:</strong> {{ report.crawl_date or "N/A" }}</p>
                </div>
                <div class="card">
                  <p><strong>Critical:</strong> {{ report.summary.issues_by_severity.critical }}</p>
                  <p><strong>High:</strong> {{ report.summary.issues_by_severity.high }}</p>
                  <p><strong>Medium:</strong> {{ report.summary.issues_by_severity.medium }}</p>
                  <p><strong>Low:</strong> {{ report.summary.issues_by_severity.low }}</p>
                </div>
              </div>
            </div>

            <div class="section">
              <h2>Charts</h2>
              <div class="grid">
                <div><img src="{{ charts.category_scores }}" alt="Category scores" /></div>
                <div><img src="{{ charts.severity }}" alt="Issues by severity" /></div>
              </div>
            </div>

            <div class="section">
              <h2>Recommendations</h2>
              <ul>
                {% for item in report.recommendations %}
                  <li><strong>{{ item.rule }}</strong> ({{ item.severity }}) - affects {{ item.pages_affected }} pages. {{ item.message }}</li>
                {% endfor %}
              </ul>
            </div>

            <div class="section" style="page-break-before: always;">
              <h2>Site Issues</h2>
              <p>One row per site-level finding — never duplicated per page.</p>
              <table>
                <thead>
                  <tr>
                    <th>Scope</th>
                    <th>Rule</th>
                    <th>Severity</th>
                    <th>Message</th>
                  </tr>
                </thead>
                <tbody>
                  {% for issue in report.site_issues %}
                    <tr>
                      <td>{{ issue.scope }}</td>
                      <td>{{ issue.rule }}</td>
                      <td>{{ issue.severity }}</td>
                      <td>{{ issue.message }}</td>
                    </tr>
                  {% endfor %}
                </tbody>
              </table>
            </div>

            <div class="section" style="page-break-before: always;">
              <h2>Page Issues</h2>
              {% for group in report.page_issues %}
                <div class="card">
                  <h3>{{ group.url }} ({{ group.issue_count }})</h3>
                  <table>
                    <thead>
                      <tr>
                        <th>Rule</th>
                        <th>Severity</th>
                        <th>Message</th>
                      </tr>
                    </thead>
                    <tbody>
                      {% for issue in group.issues %}
                        <tr>
                          <td>{{ issue.rule }}</td>
                          <td>{{ issue.severity }}</td>
                          <td>{{ issue.message }}</td>
                        </tr>
                      {% endfor %}
                    </tbody>
                  </table>
                </div>
              {% endfor %}
            </div>
          </body>
        </html>
        """


def iter_file_chunks(path: str, chunk_size: int = 65536):
    with open(path, "rb") as file_handle:
        while chunk := file_handle.read(chunk_size):
            yield chunk


def remove_file(path: str) -> None:
    try:
        os.remove(path)
    except FileNotFoundError:
        pass
